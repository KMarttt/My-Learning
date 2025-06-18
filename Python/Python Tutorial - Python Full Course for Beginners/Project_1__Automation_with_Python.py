"""For context, in this project the price in the spreadsheet
is wrong. The price should be 10% lower than the prices in the spreadsheet"""

import openpyxl as xl

# the code above renames the openpyxl as xl
# for readability

from openpyxl.chart import  BarChart, Reference
# package = openpyxl
# module = chat
# classes = Barchart, Reference (capitalized for every word)

wb = xl.load_workbook("transactions.xlsx")
# code above loads in the excel file
sheet = wb["Sheet1"]
# code above loads in the Sheet 1 (this is case sensitive)

# code below used to select cells in the sheet (for demonstration purposes)
# cell = sheet["a1"]
# cell = sheet.cell(1,1) -- (1,1) row 1, column 1
# print(cell.value) -- just checking



# print(sheet.max_row)
        # - used to indentify the number of row in the sheet

"""The code below is used to select and modify the values in the spreadsheet"""

for row in range(2,sheet.max_row + 1):
    # the values start at row 2 and max_row +1 is used to include the max row
    cell = sheet.cell(row, 3)
    # sheet.cell is used to select the cells
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    # the variable above is assigned as the name for that particular cell
    corrected_price_cell.value = corrected_price
    # which is then used to assign the cell a value = corrected price


"""The following code is used to select the values to be used on the chart"""
values = Reference(sheet,
          min_row = 2,
          max_row = sheet.max_row,
          min_col = 4,
          max_col=4)
# we want to select cell from row 2 to 4 and only from column 4 (corrected price)
# above line was modified for readability

chart = BarChart()
# assigning chart as the bar chat
chart.add_data(values)
# adding values on the chart
sheet.add_chart(chart, "e2")
# adding the coordinates where the chart will be placed
wb.save("transactions2.xlsx")


